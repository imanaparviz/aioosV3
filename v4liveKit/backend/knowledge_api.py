"""
AIOOS Knowledge Base API
RAG System - File upload, parsing, chunking, embeddings
"""

import os
import io
import csv
import asyncio
import logging
from typing import Optional, List
from uuid import UUID

from fastapi import APIRouter, UploadFile, File, HTTPException, Header, Form
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
import openai
from supabase import Client

logger = logging.getLogger("aioos-knowledge-api")

# OpenAI client for embeddings
openai.api_key = os.getenv("OPENAI_API_KEY", "")

# Maximum file size: 10MB
MAX_FILE_SIZE = 10 * 1024 * 1024

# Chunk size for text splitting (in characters)
CHUNK_SIZE = 1000
CHUNK_OVERLAP = 200


# ==================== Helper Functions ====================

def extract_text_from_txt(file_content: bytes) -> str:
    """Extract text from .txt file"""
    try:
        return file_content.decode('utf-8')
    except UnicodeDecodeError:
        # Try with different encoding
        return file_content.decode('latin-1')


def extract_text_from_csv(file_content: bytes) -> str:
    """Extract text from .csv file"""
    try:
        csv_file = io.StringIO(file_content.decode('utf-8'))
        reader = csv.reader(csv_file)
        lines = []
        for row in reader:
            lines.append(", ".join(row))
        return "\n".join(lines)
    except Exception as e:
        raise ValueError(f"Failed to parse CSV: {str(e)}")


def extract_text_from_pdf(file_content: bytes) -> str:
    """Extract text from .pdf file"""
    try:
        pdf_file = io.BytesIO(file_content)
        reader = PdfReader(pdf_file)
        text = []
        for page in reader.pages:
            text.append(page.extract_text())
        return "\n".join(text)
    except Exception as e:
        raise ValueError(f"Failed to parse PDF: {str(e)}")


def extract_text_from_docx(file_content: bytes) -> str:
    """Extract text from .docx file"""
    try:
        docx_file = io.BytesIO(file_content)
        doc = Document(docx_file)
        text = []
        for paragraph in doc.paragraphs:
            text.append(paragraph.text)
        return "\n".join(text)
    except Exception as e:
        raise ValueError(f"Failed to parse DOCX: {str(e)}")


def extract_text_from_xlsx(file_content: bytes) -> str:
    """Extract text from .xlsx file"""
    try:
        xlsx_file = io.BytesIO(file_content)
        workbook = load_workbook(xlsx_file)
        text = []

        for sheet in workbook.worksheets:
            text.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                # Filter out None values and convert to string
                row_text = ", ".join([str(cell) for cell in row if cell is not None])
                if row_text:
                    text.append(row_text)

        return "\n".join(text)
    except Exception as e:
        raise ValueError(f"Failed to parse XLSX: {str(e)}")


def extract_text_from_file(file_content: bytes, file_type: str) -> str:
    """
    Extract text from uploaded file based on file type
    """
    if file_type == "txt":
        return extract_text_from_txt(file_content)
    elif file_type == "csv":
        return extract_text_from_csv(file_content)
    elif file_type == "pdf":
        return extract_text_from_pdf(file_content)
    elif file_type == "docx":
        return extract_text_from_docx(file_content)
    elif file_type == "xlsx":
        return extract_text_from_xlsx(file_content)
    else:
        raise ValueError(f"Unsupported file type: {file_type}")


def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    """
    Split text into chunks with overlap
    This helps maintain context between chunks
    """
    if not text or len(text) == 0:
        return []

    chunks = []
    start = 0

    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]

        # Try to break at sentence boundary if possible
        if end < len(text):
            # Look for last sentence ending (.!?)
            last_period = max(chunk.rfind('.'), chunk.rfind('!'), chunk.rfind('?'))
            if last_period > chunk_size * 0.5:  # Only if we're at least halfway through
                end = start + last_period + 1
                chunk = text[start:end]

        chunks.append(chunk.strip())
        start = end - overlap

    return chunks


async def generate_embeddings(text: str) -> List[float]:
    """
    Generate embeddings using OpenAI API
    Model: text-embedding-3-small (1536 dimensions)
    """
    try:
        response = await asyncio.to_thread(
            openai.embeddings.create,
            model="text-embedding-3-small",
            input=text
        )
        return response.data[0].embedding
    except Exception as e:
        logger.error(f"Failed to generate embeddings: {e}")
        raise ValueError(f"Embeddings generation failed: {str(e)}")


# ==================== API Routes ====================

def add_knowledge_routes(app, supabase: Client):
    """
    Add knowledge base routes to FastAPI app
    """

    @app.post("/api/knowledge/upload")
    async def upload_knowledge_file(
        agent_id: str = Form(...),
        file: UploadFile = File(...),
        authorization: Optional[str] = Header(None)
    ):
        """
        Upload a file to agent's knowledge base
        Supports: .txt, .csv, .pdf, .xlsx, .docx
        """
        # Get authenticated Supabase client
        from main import get_authenticated_supabase_client
        auth_supabase = get_authenticated_supabase_client(authorization)

        try:
            # Validate file size
            file_content = await file.read()
            file_size = len(file_content)

            if file_size > MAX_FILE_SIZE:
                raise HTTPException(
                    status_code=400,
                    detail=f"File too large. Maximum size: {MAX_FILE_SIZE / 1024 / 1024}MB"
                )

            # Validate file type
            file_extension = file.filename.split(".")[-1].lower()
            supported_types = ["txt", "csv", "pdf", "xlsx", "docx"]

            if file_extension not in supported_types:
                raise HTTPException(
                    status_code=400,
                    detail=f"Unsupported file type. Supported: {', '.join(supported_types)}"
                )

            logger.info(f"📄 Uploading file: {file.filename} ({file_size} bytes)")

            # Extract text from file
            logger.info(f"📝 Extracting text from {file_extension} file...")
            extracted_text = extract_text_from_file(file_content, file_extension)

            if not extracted_text or len(extracted_text.strip()) == 0:
                raise HTTPException(status_code=400, detail="File is empty or could not extract text")

            logger.info(f"✅ Extracted {len(extracted_text)} characters")

            # Create knowledge_base record
            content_preview = extracted_text[:500]  # First 500 chars

            kb_data = {
                "agent_id": agent_id,
                "file_name": file.filename,
                "file_type": file_extension,
                "file_size_bytes": file_size,
                "content_preview": content_preview,
                "upload_status": "processing",
                "total_chunks": 0
            }

            # Insert knowledge_base record (RLS will set user_id from JWT)
            kb_response = auth_supabase.table("knowledge_base").insert(kb_data).execute()

            if not kb_response.data or len(kb_response.data) == 0:
                raise HTTPException(status_code=500, detail="Failed to create knowledge base record")

            knowledge_base_id = kb_response.data[0]["id"]
            logger.info(f"✅ Created knowledge_base record: {knowledge_base_id}")

            # Chunk the text
            logger.info(f"✂️ Chunking text...")
            chunks = chunk_text(extracted_text)
            logger.info(f"✅ Created {len(chunks)} chunks")

            # Process chunks in background
            # For now, do it synchronously (in production, use background task)
            successful_chunks = 0

            for idx, chunk in enumerate(chunks):
                try:
                    # Generate embeddings
                    embeddings = await generate_embeddings(chunk)

                    # Prepare chunk data
                    chunk_data = {
                        "knowledge_base_id": knowledge_base_id,
                        "agent_id": agent_id,
                        "chunk_index": idx,
                        "content": chunk,
                        "token_count": len(chunk.split()),  # Approximate token count
                        "embedding": embeddings,
                        "metadata": {
                            "file_name": file.filename,
                            "chunk_size": len(chunk)
                        }
                    }

                    # Insert chunk with embeddings
                    auth_supabase.table("document_chunks").insert(chunk_data).execute()
                    successful_chunks += 1

                    logger.info(f"✅ Processed chunk {idx + 1}/{len(chunks)}")

                except Exception as e:
                    logger.error(f"❌ Failed to process chunk {idx}: {e}")
                    continue

            # Update knowledge_base record with completion status
            update_data = {
                "upload_status": "completed",
                "total_chunks": successful_chunks
            }

            auth_supabase.table("knowledge_base").update(update_data).eq("id", knowledge_base_id).execute()

            logger.info(f"✅ Successfully uploaded and processed {file.filename}")

            return {
                "status": "success",
                "knowledge_base_id": knowledge_base_id,
                "file_name": file.filename,
                "file_size": file_size,
                "total_chunks": successful_chunks,
                "message": f"File uploaded and processed successfully. Created {successful_chunks} chunks."
            }

        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"❌ Failed to upload file: {e}")
            raise HTTPException(status_code=500, detail=str(e))


    @app.get("/api/knowledge/{agent_id}")
    async def get_knowledge_files(
        agent_id: str,
        authorization: Optional[str] = Header(None)
    ):
        """
        Get all knowledge base files for an agent
        """
        from main import get_authenticated_supabase_client
        auth_supabase = get_authenticated_supabase_client(authorization)

        try:
            response = auth_supabase.table("knowledge_base") \
                .select("*") \
                .eq("agent_id", agent_id) \
                .order("created_at", desc=True) \
                .execute()

            return response.data

        except Exception as e:
            logger.error(f"❌ Failed to fetch knowledge files: {e}")
            raise HTTPException(status_code=500, detail=str(e))


    @app.delete("/api/knowledge/{knowledge_base_id}")
    async def delete_knowledge_file(
        knowledge_base_id: str,
        authorization: Optional[str] = Header(None)
    ):
        """
        Delete a knowledge base file and all its chunks
        """
        from main import get_authenticated_supabase_client
        auth_supabase = get_authenticated_supabase_client(authorization)

        try:
            # Delete will cascade to document_chunks automatically
            response = auth_supabase.table("knowledge_base") \
                .delete() \
                .eq("id", knowledge_base_id) \
                .execute()

            return {
                "status": "success",
                "message": "Knowledge base file deleted successfully"
            }

        except Exception as e:
            logger.error(f"❌ Failed to delete knowledge file: {e}")
            raise HTTPException(status_code=500, detail=str(e))


    @app.post("/api/knowledge/search")
    async def search_knowledge(
        agent_id: str = Form(...),
        query: str = Form(...),
        limit: int = Form(5),
        authorization: Optional[str] = Header(None)
    ):
        """
        Search knowledge base using vector similarity
        """
        from main import get_authenticated_supabase_client
        auth_supabase = get_authenticated_supabase_client(authorization)

        try:
            # Generate embeddings for query
            query_embeddings = await generate_embeddings(query)

            # Call PostgreSQL function for vector search
            # Lowered threshold from 0.5 to 0.3 for better recall
            response = auth_supabase.rpc(
                "search_knowledge_base",
                {
                    "query_embedding": query_embeddings,
                    "agent_uuid": agent_id,
                    "match_threshold": 0.3,  # Lower threshold = more results
                    "match_count": limit
                }
            ).execute()

            return {
                "status": "success",
                "results": response.data,
                "count": len(response.data) if response.data else 0
            }

        except Exception as e:
            logger.error(f"❌ Failed to search knowledge base: {e}")
            raise HTTPException(status_code=500, detail=str(e))

        logger.info("✅ Knowledge Base API routes registered")
